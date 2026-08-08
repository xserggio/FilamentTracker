"""Filament Tracker - 3D printing filament inventory and print history.

A desktop window (pywebview) with an HTML interface and a local SQLite database.
"""

import os
import sys
import traceback

import webview

from core import (APP_DIR, DB_PATH, DRY_DAYS, SPOOL_TYPES, Store,
                  clean_url, detect_lang, guess_hex)
import catalog
import slicer
from importer import import_excel

WEB_DIR = os.path.join(APP_DIR, "web")

# pywebview 6 replaced the OPEN_DIALOG/SAVE_DIALOG constants with a FileDialog enum
_FD = getattr(webview, "FileDialog", None)
DLG_OPEN = _FD.OPEN if _FD else webview.OPEN_DIALOG
DLG_SAVE = _FD.SAVE if _FD else webview.SAVE_DIALOG
DLG_FOLDER = _FD.FOLDER if _FD else webview.FOLDER_DIALOG


def ok(data=None, **extra):
    r = {"ok": True, "data": data}
    r.update(extra)
    return r


def err(message):
    return {"ok": False, "error": str(message)}


class Api:
    def __init__(self):
        self._store = Store(DB_PATH)
        self._window = None

    # ---------- startup ----------

    def bootstrap(self):
        try:
            s = self._store
            return ok(
                {
                    "filaments": s.filaments(),
                    "prints": s.prints(),
                    "projects": s.projects(),
                    "stats": s.stats(),
                    "settings": s.get_settings(),
                    "dry_days": s.dry_days(),
                    "materials": sorted(DRY_DAYS),
                    "spool_tare": s.spool_tare(),
                    "brands": s.brands(),
                    "spool_types": list(SPOOL_TYPES),
                    "backups": s.backup_info(),
                    "empty": s.is_empty(),
                    "db_path": s.path,
                }
            )
        except Exception as e:
            traceback.print_exc()
            return err(e)

    def refresh(self, payload=None):
        payload = payload or {}
        try:
            s = self._store
            return ok(
                {
                    "filaments": s.filaments(),
                    "prints": s.prints(
                        search=payload.get("search", ""),
                        filament_id=payload.get("filament_id") or None,
                        date_from=payload.get("date_from", ""),
                        date_to=payload.get("date_to", ""),
                    ),
                    "projects": s.projects(),
                    "stats": s.stats(),
                    "settings": s.get_settings(),
                    "dry_days": s.dry_days(),
                    "materials": sorted(DRY_DAYS),
                    "spool_tare": s.spool_tare(),
                    "brands": s.brands(),
                    "spool_types": list(SPOOL_TYPES),
                    "backups": s.backup_info(),
                    "empty": s.is_empty(),
                    "db_path": s.path,
                }
            )
        except Exception as e:
            traceback.print_exc()
            return err(e)

    # ---------- filaments ----------

    def save_filament(self, data):
        try:
            if data.get("id"):
                self._store.update_filament(int(data["id"]), data)
                return ok()
            return ok({"id": self._store.add_filament(data)})
        except Exception as e:
            return err(e)

    def delete_filament(self, fid):
        try:
            self._store.delete_filament(int(fid))
            return ok()
        except Exception as e:
            return err(e)

    def set_stock(self, data):
        try:
            self._store.set_stock(int(data["id"]), int(data["stock"]))
            return ok()
        except Exception as e:
            return err(e)

    def new_roll(self, data):
        try:
            self._store.new_roll(
                int(data["id"]),
                weight=data.get("weight"),
                opened=data.get("opened"),
                brand=data.get("brand"),
                spare_id=int(data["spare_id"]) if data.get("spare_id") else None,
                from_stock=bool(data.get("from_stock", False)),
                spool_type=data.get("spool_type"),
                price=data.get("price"),
            )
            return ok()
        except Exception as e:
            return err(e)

    def add_spare(self, data):
        try:
            sid = self._store.add_spare(
                int(data["id"]), brand=data.get("brand"), weight=data.get("weight"),
                spool_type=data.get("spool_type"), price=data.get("price")
            )
            return ok({"id": sid})
        except Exception as e:
            return err(e)

    def update_spare(self, data):
        try:
            self._store.update_spare(
                int(data["spare_id"]), brand=data.get("brand"), weight=data.get("weight"),
                spool_type=data.get("spool_type"), price=data.get("price")
            )
            return ok()
        except Exception as e:
            return err(e)

    def delete_spare(self, sid):
        try:
            self._store.delete_spare(int(sid))
            return ok()
        except Exception as e:
            return err(e)

    def mark_dried(self, data):
        try:
            self._store.mark_dried(int(data["id"]), when=data.get("when"))
            return ok()
        except Exception as e:
            return err(e)

    def save_dry_days(self, data):
        try:
            self._store.set_dry_days(data or {})
            return ok()
        except Exception as e:
            return err(e)

    def adjust_roll(self, data):
        try:
            self._store.adjust_roll(int(data["id"]), float(data["remaining"]),
                                    tare=float(data.get("tare") or 0))
            return ok()
        except Exception as e:
            return err(e)

    def filament_detail(self, fid):
        """Everything about a filament for its detail sheet: rolls and prints."""
        try:
            fid = int(fid)
            fil = next((f for f in self._store.filaments() if f["id"] == fid), None)
            if fil is None:
                return err("Filament not found.")
            return ok({
                "filament": fil,
                "rolls": self._store.roll_history(fid),
                "prints": self._store.filament_prints(fid),
                "specs": catalog.specs(fil["roll_brand"], fil["material"], fil["name"]),
            })
        except Exception as e:
            traceback.print_exc()
            return err(e)

    def save_spool_tare(self, data):
        try:
            self._store.set_spool_tare(data or {})
            return ok()
        except Exception as e:
            return err(e)

    def make_backup(self):
        try:
            self._store.backup()
            return ok(self._store.backup_info())
        except Exception as e:
            return err(e)

    def open_backups(self):
        try:
            os.startfile(self._store.backup_info()["dir"])
            return ok()
        except Exception as e:
            return err(e)

    def catalog_colors(self, data):
        """The colours this manufacturer actually sells, for the colour picker."""
        try:
            return ok(catalog.colors(data.get("brand", ""), data.get("material")))
        except Exception as e:
            return err(e)

    def filament_specs(self, data):
        """Printing temperatures and density, saying where each number came from."""
        try:
            return ok(catalog.specs(data.get("brand"), data.get("material"),
                                    data.get("name", "")))
        except Exception as e:
            return err(e)

    def match_color(self, data):
        """Rank the inventory by how close each spool looks to a colour."""
        try:
            res = catalog.match_color(data.get("hex", ""), self._store.filaments(),
                                      data.get("material"))
            return ok([{"id": r["filament"]["id"], "name": r["filament"]["name"],
                        "hex": r["filament"]["hex"], "delta": r["delta"],
                        "same_material": r["same_material"]} for r in res[:8]])
        except Exception as e:
            return err(e)

    # ---------- the AMS ----------

    def ams(self):
        try:
            return ok(self._store.ams())
        except Exception as e:
            traceback.print_exc()
            return err(e)

    def set_ams_slot(self, data):
        try:
            self._store.set_ams_slot(data.get("unit"), data.get("slot"),
                                     data.get("filament_id"))
            return ok()
        except Exception as e:
            return err(e)

    # ---------- what Bambu Studio just sliced ----------

    def slices(self, data=None):
        """Slices worth offering, each with a suggested spool per filament.

        Anything already offered is filtered out by timestamp, so dismissing a
        slice makes it stay dismissed and re-slicing the same plate offers it
        again.
        """
        data = data or {}
        try:
            s = self._store
            # The setting governs the card that appears on its own, not asking
            # for the list on purpose, which is what "all" means here.
            if not data.get("all") and s.get_settings().get("slicer_watch", "1") != "1":
                return ok([])
            try:
                since = float(s.get_settings().get("slicer_seen") or 0)
            except ValueError:
                since = 0.0
            if data.get("all"):
                since = 0.0
            fils = s.filaments()
            folder = s.get_settings().get("slicer_dir", "")
            out = []
            for sl in slicer.latest_slices(limit=int(data.get("limit") or 3),
                                           since=since, custom=folder):
                for item in sl["items"]:
                    sig = slicer.signature(item)
                    item["signature"] = sig
                    item.update(slicer.candidates(item, fils, s.recall_match(sig)))
                out.append(sl)
            return ok(out)
        except Exception as e:
            traceback.print_exc()
            return err(e)

    def slicer_folder(self, data=None):
        """Which folder is being watched, whether it is there and what is in it."""
        try:
            return ok(slicer.folder_status(
                self._store.get_settings().get("slicer_dir", "")))
        except Exception as e:
            return err(e)

    def pick_slicer_folder(self):
        """Let the user point at the folder themselves when the guess is wrong."""
        try:
            paths = self._window.create_file_dialog(DLG_FOLDER)
            return ok(paths[0] if paths else None)
        except Exception as e:
            return err(e)

    def dismiss_slice(self, data):
        """Remember how far we have looked, so this slice is not offered again."""
        try:
            path = data.get("path") or ""
            stamp = os.path.getmtime(path) if path and os.path.exists(path) else 0
            self._store.set_settings({"slicer_seen": str(stamp)})
            return ok()
        except Exception as e:
            return err(e)

    def remember_matches(self, data):
        """Store the confirmations the user just gave on a slice."""
        try:
            for m in data.get("matches") or []:
                if m.get("signature") and m.get("filament_id"):
                    self._store.remember_match(m["signature"], int(m["filament_id"]))
            return ok()
        except Exception as e:
            return err(e)

    def forget_match(self, signature):
        try:
            self._store.forget_match(signature or "")
            return ok()
        except Exception as e:
            return err(e)

    def learned_matches(self):
        try:
            return ok(self._store.learned_matches())
        except Exception as e:
            return err(e)

    def guess_color(self, name):
        return ok(guess_hex(name or ""))

    # ---------- prints ----------

    def save_print(self, data):
        try:
            return ok({"id": self._store.save_print(data)})
        except Exception as e:
            return err(e)

    def delete_print(self, pid):
        try:
            self._store.delete_print(int(pid))
            return ok()
        except Exception as e:
            return err(e)

    # ---------- settings / files ----------

    def save_settings(self, data):
        try:
            self._store.set_settings(data)
            return ok()
        except Exception as e:
            return err(e)

    def pick_excel(self):
        try:
            paths = self._window.create_file_dialog(
                DLG_OPEN,
                allow_multiple=False,
                file_types=("Excel (*.xlsx;*.xlsm)", "All files (*.*)"),
            )
            return ok(paths[0] if paths else None)
        except Exception as e:
            return err(e)

    def import_excel(self, data):
        try:
            path = data.get("path")
            if not path or not os.path.exists(path):
                return err("File not found.")
            res = import_excel(self._store, path, replace=bool(data.get("replace")))
            return ok(res)
        except Exception as e:
            traceback.print_exc()
            return err(e)

    def export_excel(self):
        try:
            import openpyxl

            path = self._window.create_file_dialog(
                DLG_SAVE,
                save_filename="Filaments.xlsx",
                file_types=("Excel (*.xlsx)",),
            )
            if not path:
                return ok(None)
            if isinstance(path, (list, tuple)):
                path = path[0]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventory"
            ws.append(
                ["Filament", "Material", "Colour", "Brand", "Spool (g)", "Used (g)",
                 "Left (g)", "%", "Spares", "Opened", "Last dried"]
            )
            for f in self._store.filaments():
                ws.append([f["name"], f["material"], f["color"], f["roll_brand"],
                           f["roll_weight"], f["used"], f["remaining"], f["pct"],
                           f["stock"], f["roll_opened"], f["dried_at"]])

            ws3 = wb.create_sheet("Spares")
            ws3.append(["Filament", "Material", "Colour", "Brand", "Weight (g)"])
            for f in self._store.filaments():
                for sp in f["spares"]:
                    ws3.append([f["name"], f["material"], f["color"], sp["brand"], sp["weight"]])

            ws2 = wb.create_sheet("History")
            ws2.append(["Date", "Project", "Filament", "Grams", "Failed", "Link", "Notes"])
            for p in self._store.prints():
                for it in p["items"]:
                    ws2.append([p["date"], p["project"], it["name"], it["grams"],
                                "yes" if p["failed"] else "", p["url"], p["notes"]])

            for sheet in (ws, ws2, ws3):
                for col in sheet.columns:
                    width = max(len(str(c.value or "")) for c in col) + 2
                    sheet.column_dimensions[col[0].column_letter].width = min(38, width)
                sheet.freeze_panes = "A2"

            wb.save(path)
            return ok(path)
        except Exception as e:
            traceback.print_exc()
            return err(e)

    def open_url(self, url):
        try:
            import webbrowser

            safe = clean_url(url)
            if not safe:
                return err("That link is not valid.")
            webbrowser.open(safe)
            return ok()
        except Exception as e:
            return err(e)

    def known_materials(self):
        return ok(sorted(DRY_DAYS))

    def open_data_folder(self):
        try:
            os.startfile(os.path.dirname(self._store.path))
            return ok()
        except Exception as e:
            return err(e)


WEBVIEW2_URL = "https://developer.microsoft.com/microsoft-edge/webview2/"

# Shown only when the WebView2 runtime is missing, so it cannot go through the
# normal translation files — those live inside the window that will not open.
MISSING_RUNTIME = {
    "en": ("Filament Tracker needs the Microsoft WebView2 runtime, which is not "
           "installed on this PC.\n\nIt ships with Windows 11 and is a free "
           "download for Windows 10.\n\nOpen the download page now?"),
    "es": ("Filament Tracker necesita el runtime WebView2 de Microsoft, que no está "
           "instalado en este equipo.\n\nViene de serie en Windows 11 y es una "
           "descarga gratuita para Windows 10.\n\n¿Abrir la página de descarga?"),
    "fr": ("Filament Tracker a besoin du runtime WebView2 de Microsoft, absent de ce "
           "PC.\n\nIl est inclus dans Windows 11 et téléchargeable gratuitement pour "
           "Windows 10.\n\nOuvrir la page de téléchargement ?"),
    "de": ("Filament Tracker benötigt die Microsoft-WebView2-Laufzeit, die auf diesem "
           "PC fehlt.\n\nSie ist in Windows 11 enthalten und für Windows 10 kostenlos "
           "erhältlich.\n\nDownloadseite jetzt öffnen?"),
    "pt": ("O Filament Tracker precisa do runtime WebView2 da Microsoft, que não está "
           "instalado neste PC.\n\nVem com o Windows 11 e é uma transferência gratuita "
           "para o Windows 10.\n\nAbrir a página de transferência?"),
    "it": ("Filament Tracker richiede il runtime WebView2 di Microsoft, che non è "
           "installato su questo PC.\n\nÈ incluso in Windows 11 ed è scaricabile "
           "gratuitamente per Windows 10.\n\nAprire la pagina di download?"),
}


def has_webview2() -> bool:
    """Is the WebView2 runtime installed?

    pywebview renders through it, so without it the window never opens and the
    user just sees a stack trace. Windows 11 ships with it; Windows 10 may not.
    """
    if sys.platform != "win32":
        return True
    import winreg

    guid = "{F3017226-FE2A-4295-8BDF-00C3A9A7E4C5}"
    for hive, path in (
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\Microsoft\EdgeUpdate\Clients"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\EdgeUpdate\Clients"),
        (winreg.HKEY_CURRENT_USER, r"Software\Microsoft\EdgeUpdate\Clients"),
    ):
        try:
            with winreg.OpenKey(hive, path + "\\" + guid) as key:
                if winreg.QueryValueEx(key, "pv")[0] not in ("", "0.0.0.0"):
                    return True
        except OSError:
            continue
    return False


def warn_missing_runtime():
    """Explain the missing runtime in a plain dialog and offer the download."""
    import ctypes
    import webbrowser

    text = MISSING_RUNTIME.get(detect_lang(), MISSING_RUNTIME["en"])
    MB_YESNO_ICONWARNING = 0x34
    if ctypes.windll.user32.MessageBoxW(0, text, "Filament Tracker", MB_YESNO_ICONWARNING) == 6:
        webbrowser.open(WEBVIEW2_URL)


def main():
    if not has_webview2():
        warn_missing_runtime()
        return
    api = Api()
    window = webview.create_window(
        "Filament Tracker",
        os.path.join(WEB_DIR, "index.html"),
        js_api=api,
        width=1360,
        height=880,
        min_size=(1020, 660),
        background_color="#0f1116",
    )
    api._window = window
    debug = "--debug" in sys.argv
    icon = os.path.join(WEB_DIR, "icon.ico")
    try:
        webview.start(debug=debug, icon=icon if os.path.exists(icon) else None)
    except TypeError:
        # older pywebview backends do not accept icon=
        webview.start(debug=debug)


if __name__ == "__main__":
    main()
