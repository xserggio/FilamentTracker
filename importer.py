"""Imports a '3D Filament.xlsx' spreadsheet (Google Sheets export) into the database."""

from datetime import datetime

from core import Store, guess_hex, today


def _s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v).strip()


def _date(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    txt = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(txt[:19], fmt).date().isoformat()
        except ValueError:
            continue
    return txt[:10]


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return 0.0


def _split_name(name: str):
    """'PLA HS - Sakura pink matte' -> ('PLA HS', 'Sakura pink matte')"""
    if " - " in name:
        mat, col = name.split(" - ", 1)
        return mat.strip(), col.strip()
    return name.strip(), ""


def import_excel(store: Store, path: str, replace: bool = False) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    db = store.db

    if replace:
        db.execute("DELETE FROM print_items")
        db.execute("DELETE FROM prints")
        db.execute("DELETE FROM rolls")
        db.execute("DELETE FROM filaments")
        db.commit()

    # --- 1. Inventory sheet -> filaments --------------------------------
    known = {
        r["name"]: r["id"] for r in db.execute("SELECT id, name FROM filaments")
    }
    new_filaments = 0
    if "Inventario" in wb.sheetnames:
        ws = wb["Inventario"]
        for row in range(6, ws.max_row + 1):
            name = _s(ws.cell(row, 2).value)          # B: ID Filamento
            if not name:
                continue
            material = _s(ws.cell(row, 3).value) or _split_name(name)[0]
            color = _s(ws.cell(row, 4).value) or _split_name(name)[1]
            stock = int(_num(ws.cell(row, 9).value))  # I: Stock
            if name in known:
                db.execute("UPDATE filaments SET stock=? WHERE id=?", (stock, known[name]))
                continue
            cur = db.execute(
                "INSERT INTO filaments(name, material, color, hex, brand, stock, notes, created_at)"
                " VALUES(?,?,?,?,'',?,'',?)",
                (name, material, color, guess_hex(color), stock, today()),
            )
            known[name] = cur.lastrowid
            new_filaments += 1
    db.commit()

    # --- 2. Prints ------------------------------------------------------
    # 'Historial de Impresiones' sheet: one row per (project, filament).
    # 'Respuestas de formulario 2' sheet: up to 4 filaments per row.
    raw = []  # (date, project, filament name, grams)

    if "Historial de Impresiones" in wb.sheetnames:
        ws = wb["Historial de Impresiones"]
        for row in range(2, ws.max_row + 1):
            d = _date(ws.cell(row, 2).value)
            fil = _s(ws.cell(row, 3).value)
            g = _num(ws.cell(row, 4).value)
            proj = _s(ws.cell(row, 5).value)
            if d and fil and g > 0:
                raw.append((d, proj or "(untitled)", fil, g))

    if "Respuestas de formulario 2" in wb.sheetnames:
        ws = wb["Respuestas de formulario 2"]
        # (filament, grams) column pairs: C/D, G/H, J/K, M/N
        pairs = [(3, 4), (7, 8), (10, 11), (13, 14)]
        for row in range(2, ws.max_row + 1):
            d = _date(ws.cell(row, 2).value)
            proj = _s(ws.cell(row, 5).value)
            if not d:
                continue
            for cf, cg in pairs:
                fil = _s(ws.cell(row, cf).value)
                g = _num(ws.cell(row, cg).value)
                if fil and g > 0:
                    raw.append((d, proj or "(untitled)", fil, g))

    # Filaments that show up in the history but not in the inventory
    for _, _, fil, _ in raw:
        if fil not in known:
            mat, col = _split_name(fil)
            cur = db.execute(
                "INSERT INTO filaments(name, material, color, hex, brand, stock, notes, created_at)"
                " VALUES(?,?,?,?,'',0,'',?)",
                (fil, mat or "PLA", col, guess_hex(col), today()),
            )
            known[fil] = cur.lastrowid
            new_filaments += 1
    db.commit()

    # Group by (date, project): one print can involve several colours.
    grouped = {}
    order = []
    for d, proj, fil, g in raw:
        key = (d, proj)
        if key not in grouped:
            grouped[key] = {}
            order.append(key)
        grouped[key][fil] = grouped[key].get(fil, 0.0) + g

    existing = {
        (r["date"], r["project"])
        for r in db.execute("SELECT date, project FROM prints")
    }
    new_prints = 0
    for key in order:
        if key in existing:
            continue
        d, proj = key
        cur = db.execute(
            "INSERT INTO prints(date, project, notes, created_at) VALUES(?,?,'',?)",
            (d, proj, today()),
        )
        pid = cur.lastrowid
        for fil, g in grouped[key].items():
            db.execute(
                "INSERT INTO print_items(print_id, filament_id, grams) VALUES(?,?,?)",
                (pid, known[fil], round(g, 2)),
            )
        new_prints += 1
    db.commit()

    # --- 3. Fitted roll for each filament -------------------------------
    # The spreadsheet counted usage from the roll's opening date (a 'Rollos'
    # sheet that Google does not export). We use the first date in the history
    # so the remaining grams match what the spreadsheet showed.
    first = db.execute("SELECT MIN(date) d FROM prints").fetchone()["d"] or today()
    default_g = float(store.get_settings().get("default_spool_g", 1000))
    new_rolls = 0
    for name, fid in known.items():
        if db.execute("SELECT COUNT(*) c FROM rolls WHERE filament_id=?", (fid,)).fetchone()["c"]:
            continue
        db.execute(
            "INSERT INTO rolls(filament_id, opened_at, weight, adjust, note) "
            "VALUES(?,?,?,0,'imported from spreadsheet')",
            (fid, first, default_g),
        )
        new_rolls += 1
    db.commit()

    return {
        "filaments": new_filaments,
        "prints": new_prints,
        "rolls": new_rolls,
        "rows": len(raw),
    }
